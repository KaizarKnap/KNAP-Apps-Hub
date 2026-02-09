import io
import zipfile
import pandas as pd
import streamlit as st
import openpyxl
from copy import copy
from datetime import date as dt_date

# GEEN st.set_page_config hier (dat gebeurt in Home.py)
st.title("🧾 Self-billing per leverancier")
st.caption("Upload leveranciers-Excel(s), bereken automatisch de self-billing en exporteer het resultaat.")

# -------------------- CONFIG --------------------
CANON_COLS = [
    "Ophaaldatum", "Locatienummer", "Debiteurnummer", "Klantnaam",
    "Leverancier", "Dienst logistiek", "Uitgevoerd", "Gepland",
    "Kilogram",
    "Status", "Productomschrijving", "Afvalstroom", "Straat",
    "Huisnr", "Postcode", "Plaats", "Verantwoordelijke partij"
]

DEFAULT_PRICING_ALL = pd.DataFrame([
    {"leverancier": "Recycling-Continue", "tarieftype": "per_stop", "prijs": 4.00, "afvalstroom": ""},
    {"leverancier": "Gianluca",            "tarieftype": "per_stop", "prijs": 4.15, "afvalstroom": ""},
    {"leverancier": "Revema",              "tarieftype": "per_stop", "prijs": 4.00, "afvalstroom": ""},
    {"leverancier": "Gogogo",              "tarieftype": "per_stop", "prijs": 3.00, "afvalstroom": ""},
    {"leverancier": "Papierhandel Jansen", "tarieftype": "per_kiep", "prijs": 3.50, "afvalstroom": ""},
    {"leverancier": "Visser Assen",        "tarieftype": "per_kiep", "prijs": 4.00,  "afvalstroom": "Papier/Karton"},
    {"leverancier": "Visser Assen",        "tarieftype": "per_kiep", "prijs": 12.50, "afvalstroom": "Vertrouwelijk papier"},
    {"leverancier": "Schuman",             "tarieftype": "per_kiep", "prijs": 0.00,  "afvalstroom": ""},
    {"leverancier": "Van Bruchem",         "tarieftype": "per_kiep", "prijs": 4.00,  "afvalstroom": ""}
])

SUPPLIERS = [
    "Recycling-Continue", "Gianluca", "Revema", "Gogogo",
    "Papierhandel Jansen", "Visser Assen", "Schuman", "Van Bruchem"
]

SCHUMAN_PRICES = {
    ("240L", "Restafval"): 8.93,
    ("240L", "Papier/Karton"): 4.70,
    ("360L", "Restafval"): 12.08,
    ("360L", "Papier/Karton"): 4.70,
    ("500L", "Restafval"): 13.13,
    ("660L", "Restafval"): 16.28,
    ("660L", "Papier/Karton"): 4.70,
    ("750L", "Restafval"): 18.38,
    ("770L", "Restafval"): 18.38,
    ("770L", "Papier/Karton"): 4.70,
    ("1100L", "Restafval"): 23.63,
    ("1100L", "Papier/Karton"): 4.70,
    ("1600L", "Papier/Karton"): 4.70,
    ("1700L", "Papier/Karton"): 4.70,
    ("2400L", "Papier/Karton"): 4.70,
    ("-", "Papier/Karton"): 55.00,
}

DEFAULT_GIANLUCA_LOCS = pd.DataFrame([
    {"locatienummer": "473810001", "handelingskosten": 15.61},
    {"locatienummer": "2009100001", "handelingskosten": 15.61},
    {"locatienummer": "424980001", "handelingskosten": 15.61},
    {"locatienummer": "590930002", "handelingskosten": 15.61},
    {"locatienummer": "339960001", "handelingskosten": 15.61},
    {"locatienummer": "505660001", "handelingskosten": 15.61},
    {"locatienummer": "603760001", "handelingskosten": 15.61},
    {"locatienummer": "620640001", "handelingskosten": 15.61},
    {"locatienummer": "612540001", "handelingskosten": 15.61},
])

DEFAULT_RULES = {
    "rule_status_zero": True,
    "rule_partner_zero": True,
    "rule_client_msn_minpay": True,
    "rule_dedup_per_stop": True,
    "rule_gianluca_handlingskosten": True,
    "rule_vanbruchem_pers_92": True,
    "rule_vanbruchem_add_kg_row": True,
}

DEFAULT_KEYWORDS = ["balen", "zakken", "afzet", "pers"]

DUTCH_MONTHS = {
    1: "Januari", 2: "Februari", 3: "Maart", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Augustus", 9: "September", 10: "Oktober", 11: "November", 12: "December"
}

# -------------------- HELPERS --------------------
@st.cache_data(show_spinner=False)
def read_excel(file):
    preview = pd.read_excel(file, nrows=10, header=None)
    header_row = preview.apply(
        lambda r: r.astype(str).str.contains("Leverancier|Afvalstroom|Productomschrijving", case=False, na=False)
    ).any(axis=1)
    header_index = header_row.idxmax() if header_row.any() else 0
    df = pd.read_excel(file, header=header_index)

    df.columns = [str(c).strip().replace("\u00A0", " ") for c in df.columns]

    alias_map = {
        "# uitgevoerd": "Uitgevoerd",
        "# gepland": "Gepland",
        "aantal uitgevoerd": "Uitgevoerd",
        "aantal gepland": "Gepland",
        "uitgevoerd aantal": "Uitgevoerd",
        "gepland aantal": "Gepland"
    }
    new_cols = []
    for c in df.columns:
        key = str(c).strip().lower()
        new_cols.append(alias_map.get(key, str(c).strip()))
    df.columns = new_cols
    return df

def get_col(df, hint):
    for c in df.columns:
        if hint.lower() in str(c).lower():
            return c
    return None

def normalize_afvalstroom(v):
    v = str(v).strip().lower().replace(" ", "")
    if "papier" in v and "karton" in v:
        return "Papier/Karton"
    if "rest" in v:
        return "Restafval"
    if "vertrouw" in v:
        return "Vertrouwelijk papier"
    return v

def normalize_volume(v):
    v = str(v).strip().upper().replace(" ", "")
    if v.isdigit() and not v.endswith("L"):
        v += "L"
    return v

def normalize_volume_any(v):
    s = str(v).strip().upper().replace(" ", "")
    if s.isdigit():
        return s + "L"
    return s.replace("M³", "M3")

def normalize_loc(l):
    if l is None:
        return ""
    s = str(l).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def units_from_row(row, tarieftype):
    val = row.get("Uitgevoerd", 0)
    if tarieftype == "per_kiep":
        try:
            return int(val)
        except Exception:
            return 1 if str(val).strip() else 0
    return 1 if str(val).strip() else 0

def match_price(row, pricing_df, supplier):
    afst = str(row.get("Afvalstroom", "")).strip()
    df = pricing_df[pricing_df["leverancier"].str.lower().str.contains(supplier.lower())]

    if supplier.lower() == "visser assen":
        for _, r in df.iterrows():
            if str(r.get("afvalstroom", "")).lower() == str(afst).lower():
                return {"tarieftype": "per_kiep", "prijs": float(r["prijs"])}

    if not df.empty:
        r = df.iloc[0]
        return {"tarieftype": r["tarieftype"], "prijs": float(r["prijs"])}

    return {"tarieftype": "per_stop", "prijs": 0.0}

def get_kg_value(row, df):
    candidates = []
    for c in df.columns:
        cl = str(c).lower().strip()
        if "gewicht" in cl or cl == "kg" or "kilo" in cl:
            candidates.append(c)
    for c in candidates:
        v = row.get(c, None)
        if pd.notna(v) and str(v).strip() != "":
            return v
    return None

def export_excel_bytes(out_df: pd.DataFrame) -> bytes:
    out_export = out_df.copy()
    out_export.columns = [c.upper() for c in out_export.columns]

    if "OPHAALDATUM" in out_export.columns:
        out_export["OPHAALDATUM"] = pd.to_datetime(out_export["OPHAALDATUM"], errors="coerce").dt.strftime("%d-%m-%Y")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        out_export.to_excel(writer, index=False, sheet_name="SelfBilling")
        worksheet = writer.sheets["SelfBilling"]

        for idx, col in enumerate(out_export.columns):
            try:
                max_len = min(max(out_export[col].astype(str).map(len).max(), len(col)) + 2, 30)
            except Exception:
                max_len = min(len(col) + 2, 30)
            worksheet.set_column(idx, idx, max_len)

        if "BEDRAG" in out_export.columns:
            last_row = len(out_export) + 2
            bedrag_col = out_export.columns.get_loc("BEDRAG")
            label_col_idx = max(bedrag_col - 1, 0)

            def excel_col(n):
                s = ""
                n += 1
                while n:
                    n, r = divmod(n - 1, 26)
                    s = chr(65 + r) + s
                return s

            label_cell = f"{excel_col(label_col_idx)}{last_row}"
            bedrag_cell = f"{excel_col(bedrag_col)}{last_row}"
            bedrag_range = f"{excel_col(bedrag_col)}2:{excel_col(bedrag_col)}{last_row-1}"

            worksheet.write(label_cell, "Totaal te ontvangen bedrag")
            worksheet.write_array_formula(f"{bedrag_cell}:{bedrag_cell}", f"=SOM({bedrag_range})")

    buf.seek(0)
    return buf.getvalue()

def apply_product_filter(data: pd.DataFrame, product_col: str, allowed_products: set | None) -> pd.DataFrame:
    if allowed_products is None or not product_col:
        return data
    return data[data[product_col].astype(str).isin(allowed_products)].copy()

def compute_allowed_products_toggles(df_sup: pd.DataFrame, supplier_name: str, keywords: list[str]) -> set | None:
    product_col = next((c for c in df_sup.columns if "productomschrijving" in str(c).lower()), None)
    if product_col is None:
        st.info("Geen kolom ‘Productomschrijving’ gevonden. Er wordt niet gefilterd.")
        return None

    products_all = sorted(df_sup[product_col].dropna().astype(str).unique().tolist())
    if not products_all:
        st.info("Geen productomschrijvingen gevonden. Er wordt niet gefilterd.")
        return None

    impacted = [p for p in products_all if any(k in p.lower() for k in keywords)]
    unaffected = [p for p in products_all if p not in impacted]

    st.write(f"🔎 Gevonden **{len(impacted)}** met trefwoord en **{len(unaffected)}** zonder trefwoord.")
    st.caption("Alles zonder trefwoord gaat altijd mee. Zet hieronder alleen trefwoord-regels aan/uit.")

    active_impacted = []
    if impacted:
        cols = st.columns(2)
        for i, prod in enumerate(impacted):
            key = f"tgl_{supplier_name}_{hash(prod)}"
            with cols[i % 2]:
                if st.toggle(prod, value=True, key=key):
                    active_impacted.append(prod)

    allowed = set(unaffected) | set(active_impacted)
    excluded = set(impacted) - set(active_impacted)

    st.success(f"✅ Meegenomen: {len(allowed)} productomschrijvingen (waarvan {len(active_impacted)} met trefwoord).")
    if excluded:
        st.warning(f"🚫 Uitgesloten (trefwoord): {len(excluded)}")

    return allowed

def detect_month_year_from_df(df: pd.DataFrame) -> tuple[int | None, int | None]:
    # Pak eerste geldige datum uit OPHAALDATUM/Ophaaldatum kolom
    col = None
    for c in df.columns:
        if str(c).strip().lower() in ("ophaaldatum", "ophaal datum", "ophaal_datum"):
            col = c
            break
    if col is None:
        return None, None
    ser = pd.to_datetime(df[col], errors="coerce")
    ser = ser.dropna()
    if ser.empty:
        return None, None
    d = ser.iloc[0].to_pydatetime().date()
    return d.month, d.year

def process_supplier(
    data_all: pd.DataFrame,
    supplier_name: str,
    pricing_df: pd.DataFrame,
    rules: dict,
    gianluca_exceptions: pd.DataFrame,
    allowed_products: set | None
) -> pd.DataFrame:
    supplier = supplier_name.lower()

    supplier_col = get_col(data_all, "leverancier")
    if not supplier_col:
        return pd.DataFrame()

    data = data_all[data_all[supplier_col].astype(str).str.lower().str.contains(supplier)].copy()
    if data.empty:
        return pd.DataFrame()

    product_col = next((c for c in data.columns if "productomschrijving" in str(c).lower()), None)
    data = apply_product_filter(data, product_col, allowed_products)
    if data.empty:
        return pd.DataFrame()

    loc_dict = {}
    if supplier == "gianluca" and rules.get("rule_gianluca_handlingskosten", True):
        loc_dict = {
            normalize_loc(r["locatienummer"]): float(r["handelingskosten"])
            for _, r in gianluca_exceptions.iterrows()
            if pd.notna(r.get("locatienummer", None))
        }

    results = []
    seen_per_stop = set()

    for _, row in data.iterrows():
        product_col_local = next((c for c in data.columns if "productomschrijving" in str(c).lower()), None)
        prod_txt = str(row.get(product_col_local, "")) if product_col_local else ""
        vol_col = get_col(data, "volume")
        afst_col = get_col(data, "afvalstroom")
        vol_any = normalize_volume_any(row.get(vol_col, "")) if vol_col else ""
        afst_norm = normalize_afvalstroom(row.get(afst_col, "")) if afst_col else ""

        is_vanbruchem_pers_23_pk = (
            supplier == "van bruchem"
            and ("pers" in prod_txt.lower())
            and (vol_any == "23M3")
            and (afst_norm == "Papier/Karton")
        )

        # ---- prijsberekening ----
        if supplier == "schuman":
            ttype = "per_kiep"
            volume = normalize_volume(row.get(vol_col, "")) if vol_col else ""
            afst = normalize_afvalstroom(row.get(afst_col, "")) if afst_col else ""
            prijs = float(SCHUMAN_PRICES.get((volume, afst), 0.0))
            qty = units_from_row(row, ttype)
            bedrag = prijs * qty

        elif supplier == "gianluca":
            info = match_price(row, pricing_df, supplier_name)
            prijs = float(info["prijs"])
            qty = units_from_row(row, info["tarieftype"])
            bedrag = prijs if qty > 0 else 0.0

            if rules.get("rule_gianluca_handlingskosten", True):
                loc = normalize_loc(row.get("Locatienummer", ""))
                status_txt = str(row.get("Status", "")).strip().lower()
                if status_txt == "voltooid" and loc in loc_dict:
                    bedrag += loc_dict[loc]

        elif supplier == "visser assen":
            info = match_price(row, pricing_df, supplier_name)
            prijs = float(info["prijs"])
            qty = units_from_row(row, info["tarieftype"])
            bedrag = prijs * qty

        elif supplier == "van bruchem":
            info = match_price(row, pricing_df, supplier_name)
            prijs = float(info["prijs"])
            qty = units_from_row(row, info["tarieftype"])
            bedrag = prijs if info["tarieftype"] == "per_stop" and qty > 0 else prijs * qty

            if rules.get("rule_vanbruchem_pers_92", True) and is_vanbruchem_pers_23_pk and qty > 0:
                prijs = 92.0
                bedrag = 92.0

        else:
            info = match_price(row, pricing_df, supplier_name)
            prijs = float(info["prijs"])
            qty = units_from_row(row, info["tarieftype"])
            bedrag = prijs if info["tarieftype"] == "per_stop" and qty > 0 else prijs * qty

        # ---- status/verantwoordelijke ----
        verantwoordelijke = str(row.get("Verantwoordelijke partij", "")).strip().lower()
        status = str(row.get("Status", "")).strip().lower()

        if rules.get("rule_status_zero", True) and status in ("gepland", "geannuleerd", "discussie"):
            bedrag = 0.0
        elif rules.get("rule_partner_zero", True) and verantwoordelijke == "partner":
            bedrag = 0.0
        elif rules.get("rule_client_msn_minpay", True) and verantwoordelijke in ("client", "msn") and bedrag == 0 and prijs > 0:
            bedrag = prijs

        # ---- dedup per_stop ----
        if rules.get("rule_dedup_per_stop", True) and supplier in ("recycling-continue", "gianluca", "revema", "gogogo"):
            loc_key = normalize_loc(row.get("Locatienummer", ""))

            ophaal_raw = row.get("Ophaaldatum", None)
            ophaal_dt = pd.to_datetime(ophaal_raw, errors="coerce")
            dag_key = ophaal_dt.date().isoformat() if pd.notna(ophaal_dt) else str(ophaal_raw).strip()

            stop_key = (dag_key, loc_key)

            if bedrag > 0 and prijs > 0 and abs(bedrag - prijs) < 1e-9:
                if stop_key in seen_per_stop:
                    bedrag = 0.0
                else:
                    seen_per_stop.add(stop_key)

        # ---- base row (Kilogram altijd leeg) ----
        base_row = {
            **{c: row.get(c, None) for c in CANON_COLS if c in data.columns},
            "Kilogram": None,
            "Prijs per stuk": prijs,
            "Bedrag": bedrag
        }
        results.append(base_row)

        # ---- extra kg-regel onder pers (alleen kg kolom gevuld) ----
        if supplier == "van bruchem" and rules.get("rule_vanbruchem_add_kg_row", True) and is_vanbruchem_pers_23_pk:
            kg_row = {
                **{c: row.get(c, None) for c in CANON_COLS if c in data.columns},
                "Kilogram": get_kg_value(row, data),
                "Productomschrijving": "Kilogrammen opgehaald met pers (23m3)",
                "Prijs per stuk": 0.0,
                "Bedrag": 0.0
            }
            results.append(kg_row)

    return pd.DataFrame(results)

# -------------------- INLEESBESTAND HELPERS --------------------
def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """Kopieer cel-stijl (incl. number formats) van src_row naar dst_row."""
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.comment = None

def build_inleesbestand_from_template(template_bytes: bytes, rows: list[dict]) -> bytes:
    """
    Bouwt een nieuw Excel bestand op basis van template, zonder headers/stijl te wijzigen.
    We vullen alleen waarden in bestaande kolommen.
    """
    buf = io.BytesIO(template_bytes)
    wb = openpyxl.load_workbook(buf)
    ws = wb[wb.sheetnames[0]]

    max_col = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    header_map = {str(h).strip(): idx + 1 for idx, h in enumerate(headers) if h is not None}

    # template heeft doorgaans 1 voorbeeldregel op rij 2 -> gebruik die stijl als “basis”
    style_row = 2 if ws.max_row >= 2 else 1
    start_row = 2  # overschrijf sample-rij als die er is
    # Als er in jouw template al meerdere voorbeeldregels staan: dan gaan we append vanaf laatste+1
    if ws.max_row > 2:
        start_row = ws.max_row + 1

    for i, r in enumerate(rows):
        target_row = start_row + i

        # stijl kopiëren
        if target_row != style_row:
            copy_row_style(ws, style_row, target_row, max_col)

        # waarden invullen
        for k, v in r.items():
            if k not in header_map:
                continue  # mag niks aan koppen veranderen
            col_idx = header_map[k]
            ws.cell(row=target_row, column=col_idx).value = v

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# -------------------- UI: MODE --------------------
mode = st.radio(
    "Kies modus:",
    ["1 leverancier", "Alle leveranciers tegelijk (per leverancier Excel)"],
    index=0
)
process_all = (mode != "1 leverancier")

# -------------------- UI: KEYWORDS --------------------
st.subheader("Trefwoorden (voor productomschrijvingen)")
kw_text = st.text_input(
    "Trefwoorden (komma-gescheiden)",
    value=", ".join(DEFAULT_KEYWORDS),
    help="Alleen productomschrijvingen die één van deze woorden bevatten kun je per leverancier aan/uit zetten."
)
KEYWORDS = [k.strip().lower() for k in kw_text.split(",") if k.strip()]

# -------------------- UI: INLEESBESTAND OPTIE --------------------
st.subheader("📥 Inleesbestand (Order inlezen)")
make_inlees = st.checkbox("Maak ook een inleesbestand (1 regel per leverancier)", value=False)

template_file = None
inlees_date = None
loc_map_editor = None

if make_inlees:
    template_file = st.file_uploader(
        "Upload Template orders inlezen.xlsx (koppen/stijl mogen niet wijzigen)",
        type=["xlsx"],
        accept_multiple_files=False,
        key="inlees_template"
    )
    inlees_date = st.date_input("Uitvoerdatum (voor inleesbestand)", value=dt_date.today())

    st.caption("Vul per leverancier de Locatiecode in (verplicht voor inleesbestand).")
    loc_map_default = pd.DataFrame([{"Leverancier": s, "Locatiecode": ""} for s in SUPPLIERS])
    loc_map_editor = st.data_editor(loc_map_default, use_container_width=True, num_rows="fixed", key="loc_map_editor")

# -------------------- UI: RULES + PRICING --------------------
st.subheader("Instellingen per leverancier")

pricing_by_supplier = {}
rules_by_supplier = {}
gianluca_exceptions = DEFAULT_GIANLUCA_LOCS.copy()

def supplier_pricing_default(s: str) -> pd.DataFrame:
    return DEFAULT_PRICING_ALL[DEFAULT_PRICING_ALL["leverancier"].str.lower().str.contains(s.lower())].reset_index(drop=True)

def supplier_rules_editor(s: str, key_prefix: str) -> dict:
    with st.expander(f"Regels: {s}", expanded=False):
        r = DEFAULT_RULES.copy()

        r["rule_status_zero"] = st.checkbox("Status (Gepland/Geannuleerd/Discussie) => bedrag 0", value=r["rule_status_zero"], key=f"{key_prefix}_status")
        r["rule_partner_zero"] = st.checkbox("Verantwoordelijke partij = Partner => bedrag 0", value=r["rule_partner_zero"], key=f"{key_prefix}_partner")
        r["rule_client_msn_minpay"] = st.checkbox("Client/MSN: als bedrag 0 maar prijs>0 => alsnog prijs", value=r["rule_client_msn_minpay"], key=f"{key_prefix}_clientmsn")
        r["rule_dedup_per_stop"] = st.checkbox("Deduplicatie per_stop (zelfde dag + locatienummer => niet 2x stopfee)", value=r["rule_dedup_per_stop"], key=f"{key_prefix}_dedup")

        if s.lower() == "gianluca":
            r["rule_gianluca_handlingskosten"] = st.checkbox("Gianluca: handelingskosten op uitzonderingslocaties (alleen bij VOLTOOID)", value=r["rule_gianluca_handlingskosten"], key=f"{key_prefix}_gianluca_hk")

        if s.lower() == "van bruchem":
            r["rule_vanbruchem_pers_92"] = st.checkbox("Van Bruchem: pers 23m3 Papier/Karton = €92 (alleen bij uitgevoerd)", value=r["rule_vanbruchem_pers_92"], key=f"{key_prefix}_vb_pers92")
            r["rule_vanbruchem_add_kg_row"] = st.checkbox("Van Bruchem: extra kilogram-regel onder persregel (Bedrag = 0)", value=r["rule_vanbruchem_add_kg_row"], key=f"{key_prefix}_vb_kgrow")

        return r

if process_all:
    for s in SUPPLIERS:
        with st.expander(f"Tarieven: {s}", expanded=False):
            pricing_by_supplier[s] = st.data_editor(
                supplier_pricing_default(s),
                use_container_width=True,
                num_rows="dynamic",
                key=f"pricing_{s}"
            )
        rules_by_supplier[s] = supplier_rules_editor(s, key_prefix=f"rules_{s}")

        if s.lower() == "schuman":
            with st.expander("Prijstabel Schuman", expanded=False):
                st.dataframe(
                    pd.DataFrame([{"Volume": v[0], "Afvalstroom": v[1], "Prijs (€)": p} for v, p in SCHUMAN_PRICES.items()]),
                    use_container_width=True,
                )

    with st.expander("Gianluca: uitzonderingen / handelingskosten", expanded=False):
        st.markdown("Locatienummers hieronder krijgen handelingskosten (alleen bij status VOLTOOID).")
        gianluca_exceptions = st.data_editor(
            DEFAULT_GIANLUCA_LOCS, use_container_width=True, num_rows="dynamic", key="gianluca_ex_all"
        )

else:
    selected_supplier = st.selectbox("Kies leverancier:", SUPPLIERS)
    st.info(f"Berekening en export gelden alleen voor **{selected_supplier}**.")

    with st.expander("Tarieven (geselecteerde leverancier)", expanded=True):
        pricing_by_supplier[selected_supplier] = st.data_editor(
            supplier_pricing_default(selected_supplier),
            use_container_width=True,
            num_rows="dynamic",
            key="pricing_single"
        )

    rules_by_supplier[selected_supplier] = supplier_rules_editor(selected_supplier, key_prefix="rules_single")

    if selected_supplier.lower() == "schuman":
        with st.expander("Prijstabel Schuman", expanded=False):
            st.dataframe(
                pd.DataFrame([{"Volume": v[0], "Afvalstroom": v[1], "Prijs (€)": p} for v, p in SCHUMAN_PRICES.items()]),
                use_container_width=True,
            )

    if selected_supplier.lower() == "gianluca":
        with st.expander("Gianluca: uitzonderingen / handelingskosten", expanded=False):
            gianluca_exceptions = st.data_editor(
                DEFAULT_GIANLUCA_LOCS, use_container_width=True, num_rows="dynamic", key="gianluca_ex_single"
            )
    else:
        gianluca_exceptions = pd.DataFrame(columns=["locatienummer", "handelingskosten"])

# -------------------- UPLOAD --------------------
st.subheader("Upload leveranciers-Excel(s)")
files = st.file_uploader("Selecteer Excel-bestanden", type=["xlsx", "xls"], accept_multiple_files=True)

if not files:
    st.info("Upload één of meer Excel-bestanden om te berekenen en te exporteren.")
    st.stop()

frames = [read_excel(f) for f in files]
data_all = pd.concat(frames, ignore_index=True)

supplier_col = get_col(data_all, "leverancier")
if not supplier_col:
    st.error("Geen kolom gevonden die ‘Leverancier’ bevat. Controleer je Excel.")
    st.stop()

# -------------------- PRODUCT FILTERS: PER LEVERANCIER (TOGGLES) --------------------
allowed_products_by_supplier: dict[str, set | None] = {}

if process_all:
    st.subheader("🧩 Productomschrijvingen per leverancier (bulk)")
    st.caption("Per leverancier kun je alleen de productomschrijvingen mét trefwoord aan/uit zetten.")

    present_suppliers = [s for s in SUPPLIERS if data_all[supplier_col].astype(str).str.lower().str.contains(s.lower()).any()]
    if not present_suppliers:
        st.warning("Geen bekende leveranciers gevonden in dit bestand.")
        st.stop()

    for s in present_suppliers:
        df_sup = data_all[data_all[supplier_col].astype(str).str.lower().str.contains(s.lower())].copy()
        with st.expander(f"Productomschrijvingen: {s}", expanded=False):
            allowed_products_by_supplier[s] = compute_allowed_products_toggles(df_sup, supplier_name=s, keywords=KEYWORDS)

else:
    st.subheader("🧩 Productomschrijvingen (geselecteerde leverancier)")
    df_sup = data_all[data_all[supplier_col].astype(str).str.lower().str.contains(selected_supplier.lower())].copy()
    allowed_products_by_supplier[selected_supplier] = compute_allowed_products_toggles(df_sup, supplier_name=selected_supplier, keywords=KEYWORDS)

# -------------------- RUN --------------------
if process_all:
    st.subheader("Resultaten per leverancier")

    zip_buf = io.BytesIO()
    inlees_rows = []
    loc_map = {}

    if make_inlees and loc_map_editor is not None:
        for _, r in loc_map_editor.iterrows():
            loc_map[str(r["Leverancier"]).strip()] = str(r["Locatiecode"]).strip()

    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for s in present_suppliers:
            out_df = process_supplier(
                data_all=data_all,
                supplier_name=s,
                pricing_df=pricing_by_supplier.get(s, supplier_pricing_default(s)),
                rules=rules_by_supplier.get(s, DEFAULT_RULES),
                gianluca_exceptions=gianluca_exceptions if s.lower() == "gianluca" else pd.DataFrame(columns=["locatienummer", "handelingskosten"]),
                allowed_products=allowed_products_by_supplier.get(s, None)
            )

            if out_df.empty:
                st.markdown(f"**{s}** — geen regels na filtering.")
                continue

            st.markdown(f"**{s}** — {len(out_df)} regels")
            st.dataframe(out_df.head(20), use_container_width=True)

            # Selfbilling export
            xbytes = export_excel_bytes(out_df)
            fname = f"selfbilling_{s.lower().replace(' ', '_')}.xlsx"
            zf.writestr(fname, xbytes)

            # Inlees regel verzamelen (1 per leverancier)
            if make_inlees:
                total = float(pd.to_numeric(out_df.get("Bedrag", 0), errors="coerce").fillna(0).sum())
                # totaal moet NEGATIEF in Kenmerk
                kenmerk_val = -round(total, 2)

                m, y = detect_month_year_from_df(out_df)
                if m is None or y is None:
                    m = inlees_date.month
                    y = inlees_date.year

                boekstuk = f"Selfbilling maand {DUTCH_MONTHS.get(m, str(m))} {y}"

                loc_code = loc_map.get(s, "")
                inlees_rows.append({
                    "Locatiecode": loc_code,
                    "Artikelcode": 900100,
                    "Aantal": 1,
                    "Uitvoerdatum": inlees_date.strftime("%d-%m-%Y"),
                    "Dienst": "Administratief",
                    "Dienst_Factureren": "Standaard",
                    "Kenmerk": kenmerk_val,
                    "Boekstuknummer": boekstuk,
                    "Lev_AddresNumber": ""  # leeg laten
                })

        # Inleesbestand maken + in ZIP stoppen
        if make_inlees:
            if template_file is None:
                st.error("Je hebt ‘Maak inleesbestand’ aangezet, maar nog geen template geüpload.")
            else:
                # check locatiecodes ingevuld
                missing = [r for r in inlees_rows if not str(r.get("Locatiecode", "")).strip()]
                if missing:
                    st.error("Inleesbestand: niet alle Locatiecode waarden zijn ingevuld. Vul ze in bij ‘Locatiecode per leverancier’.")
                else:
                    template_bytes = template_file.getvalue()
                    inlees_bytes = build_inleesbestand_from_template(template_bytes, inlees_rows)
                    zf.writestr("orders_inlezen_selfbilling.xlsx", inlees_bytes)

    zip_buf.seek(0)
    st.download_button(
        label="💾 Download alle self-billings (ZIP)",
        data=zip_buf.getvalue(),
        file_name="selfbilling_per_leverancier.zip",
        mime="application/zip"
    )

    # Extra: losse download van inleesbestand (handig, zonder ZIP)
    if make_inlees and template_file is not None and inlees_rows:
        # alleen aanbieden als alle locatiecodes ingevuld
        if all(str(r.get("Locatiecode", "")).strip() for r in inlees_rows):
            inlees_bytes_single = build_inleesbestand_from_template(template_file.getvalue(), inlees_rows)
            st.download_button(
                label="📥 Download inleesbestand (orders_inlezen_selfbilling.xlsx)",
                data=inlees_bytes_single,
                file_name="orders_inlezen_selfbilling.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

else:
    out_df = process_supplier(
        data_all=data_all,
        supplier_name=selected_supplier,
        pricing_df=pricing_by_supplier[selected_supplier],
        rules=rules_by_supplier[selected_supplier],
        gianluca_exceptions=gianluca_exceptions if selected_supplier.lower() == "gianluca" else pd.DataFrame(columns=["locatienummer", "handelingskosten"]),
        allowed_products=allowed_products_by_supplier.get(selected_supplier, None)
    )

    st.subheader("Bekijk en exporteer resultaat")
    st.dataframe(out_df.head(40), use_container_width=True)

    xbytes = export_excel_bytes(out_df)
    st.download_button(
        label=f"💾 Download self-billing ({selected_supplier}).xlsx",
        data=xbytes,
        file_name=f"selfbilling_{selected_supplier.lower().replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Inleesbestand single-supplier
    if make_inlees:
        if template_file is None:
            st.error("Je hebt ‘Maak inleesbestand’ aangezet, maar nog geen template geüpload.")
        else:
            loc_code = ""
            if loc_map_editor is not None:
                row = loc_map_editor[loc_map_editor["Leverancier"].astype(str).str.strip().eq(selected_supplier)]
                if not row.empty:
                    loc_code = str(row.iloc[0]["Locatiecode"]).strip()

            if not loc_code:
                st.error("Inleesbestand: Locatiecode is leeg. Vul deze in bij ‘Locatiecode per leverancier’.")
            else:
                total = float(pd.to_numeric(out_df.get("Bedrag", 0), errors="coerce").fillna(0).sum())
                kenmerk_val = -round(total, 2)

                m, y = detect_month_year_from_df(out_df)
                if m is None or y is None:
                    m = inlees_date.month
                    y = inlees_date.year

                boekstuk = f"Selfbilling maand {DUTCH_MONTHS.get(m, str(m))} {y}"

                inlees_rows = [{
                    "Locatiecode": loc_code,
                    "Artikelcode": 900100,
                    "Aantal": 1,
                    "Uitvoerdatum": inlees_date.strftime("%d-%m-%Y"),
                    "Dienst": "Administratief",
                    "Dienst_Factureren": "Standaard",
                    "Kenmerk": kenmerk_val,
                    "Boekstuknummer": boekstuk,
                    "Lev_AddresNumber": ""
                }]

                inlees_bytes = build_inleesbestand_from_template(template_file.getvalue(), inlees_rows)
                st.download_button(
                    label="📥 Download inleesbestand (orders_inlezen_selfbilling.xlsx)",
                    data=inlees_bytes,
                    file_name="orders_inlezen_selfbilling.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
